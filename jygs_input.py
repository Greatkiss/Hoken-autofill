from input_reconfirm import jygs_name_input, jygs_num_input, reconfirm, reconfirm_num
import openpyxl as xl

class jygs:
    def __init__(self):
        self.name = ""
        self.kohonum = [""]*3
        self.seirinum = [""]*2
        self.shahonum = ""
        self.postnum=[""]*3
        self.address = ""
        self.simei = self.name
        self.phone = [""]*3

    def load_info(self):
        print("事業所の列数を入力(jygs_info.xlxsを参照)")
        n = int(input())
        book = xl.load_workbook("jygs_info.xlsx")
        sheet = book["Sheet1"]
        self.name = sheet.cell(row=n,column=1).value
        self.postnum[0] = str(sheet.cell(row=n,column=2).value)
        self.postnum[1] = str(sheet.cell(row=n,column=3).value)
        self.address = str(sheet.cell(row=n,column=4).value)
        self.simei = str(sheet.cell(row=n,column=5).value)
        self.phone[0] = str(sheet.cell(row=n,column=6).value)
        self.phone[1] = str(sheet.cell(row=n,column=7).value)
        self.phone[2] = str(sheet.cell(row=n,column=8).value)
        self.seirinum[0] = str(sheet.cell(row=n,column=9).value)
        self.seirinum[1] = str(sheet.cell(row=n,column=10).value)
        self.shahonum = str(sheet.cell(row=n,column=11).value)
        self.kohonum[0] = str(sheet.cell(row=n,column=12).value)
        self.kohonum[1] = str(sheet.cell(row=n,column=13).value)
        self.kohonum[2] = str(sheet.cell(row=n,column=14).value)