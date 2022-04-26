#from NotUsing.input_reconfirm import jygs_name_input, jygs_num_input, reconfirm, reconfirm_num, kysh_name_input, reconfirm_name, kysh_num_input, kysh_reconfirm_num
import openpyxl as xl

class kysh:
    def __init__(self):
        self.hiho = [""]*3
        self.kubun = ""
        self.name_kana = [""]*2
        self.name_kanji = [""]*2
        self.sex = ""
        self.birth = [""]*4
        self.getdate = [""]*4
        self.tin = ""
        self.carfare = ""
        self.bikou = ""
        self.kojin = ""

    def load_info(self,n):
        book = xl.load_workbook("kysh_info.xlsx")
        sheet = book["Sheet1"]
        self.hiho = cell_to_list(sheet.cell(row=n,column=3).value)
        self.kojin = str(sheet.cell(row=n,column=4).value)
        self.kubun = str(sheet.cell(row=n,column=5).value)
        self.sex = str(sheet.cell(row=n,column=6).value)
        self.birth = cell_to_list(str(sheet.cell(row=n,column=7).value))
        self.getdate = cell_to_list(str(sheet.cell(row=n,column=8).value))
        self.tin = str(sheet.cell(row=n,column=9).value)
        self.carfare = str(sheet.cell(row=n,column=10).value)
        self.bikou = str(sheet.cell(row=n,column=11).value)
        self.name_kana = cell_to_name(str(sheet.cell(row=n,column=2).value))
        self.name_kanji = cell_to_name(str(sheet.cell(row=n,column=1).value))


def cell_to_list(cell_data):
    if(str(cell_data).find('-')):
        l = str(cell_data).split('-')
    elif(cell_data.find(' ')):
        l = str(cell_data).split(' ')
    return l


def cell_to_name(cell_data):
    if(str(cell_data).find('-')>0):
        l = str(cell_data).split('-')
    elif(str(cell_data).find(' ')>0):
        l = str(cell_data).split(' ')
    elif(str(cell_data).find('　')>0):
        l = str(cell_data).split('　')
    else:
        print("error")
        return -1
    
    if(len(l) < 3):
        return l
    else:
        l_foreign = [""]*2
        l_foreign[0] = l[0]
        l.pop(0)
        l_foreign[1] = ' '.join(l)
        return l_foreign